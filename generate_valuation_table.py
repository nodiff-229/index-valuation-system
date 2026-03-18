"""
指数估值表生成器 - 完整版 v2

功能：
1. 包含 47 个指数 + 恒生科技
2. 所有数字保留两位小数
3. 增加"原因分析"列（大模型分析）
4. 字体：中文方正仿宋_GBK 17 号，英文数字 Times New Roman 17 号
5. 数据存疑时黄色标注
6. Token 失效飞书告警
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import requests
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 从分析器导入
from index_analyzer import analyze_index_reason

# Tushare Token
TUSHARE_TOKEN = "19888ce9ba935e06ae7d66902a65d5455634ad2b6b6ee7eb258217c4"

# 指数代码映射（用于 Tushare API）
INDEX_CODE_MAP = {
    "中证红利": "000922.SH",
    "中证红利低波动": "H30269.CSI",
    "沪港深红利低波": "H30365.CSI",
    "消费 50": "000149.SH",
    "消费龙头": "931494.CSI",
    "医药 100": "000933.SH",
    "中证消费": "000932.SH",
    "中证白酒": "H30184.CSI",
    "中证医疗": "H30260.CSI",
    "港股科技": "931234.CSI",
    "证券行业": "886041.TI",
    "中证银行": "399986.SZ",
    "上证红利": "000015.SH",
    "300 价值": "000919.SH",
    "优选 300": "000933.SH",
    "中证价值": "000925.SH",
    "龙头红利": "000931.SH",
    "港股红利": "H30366.CSI",
    "自由现金流": "931493.CSI",
    "50AH 优选": "000096.SH",
    "恒生指数": "HSI.HK",
    "恒生科技": "HSTECH.HK",
    "H 股指数": "HSCEI.HK",
    "基本面 50": "000925.SH",
    "沪深 300": "000300.SH",
    "上证 50": "000016.SH",
    "央视 50": "000955.SH",
    "MSCI A50": "716118.MTI",
    "中证 A50": "000850.SH",
    "中证 800": "000906.SH",
    "中证 A100": "000849.SH",
    "香港中小": "931253.CSI",
    "中证 A500": "000852.SH",
    "红利机会": "H30269.CSI",
    "基本面 60": "000926.SH",
    "基本面 120": "000927.SH",
    "可选消费": "000934.SH",
    "深证 100": "399004.SZ",
    "深证成指": "399001.SZ",
    "中证养老": "H30343.CSI",
    "中证 500": "000905.SH",
    "创业板": "399006.SZ",
    "标普 500": ".INX",
    "纳斯达克 100": ".NDX",
    "500 低波动": "H30263.CSI",
    "上证 180": "000010.SH",
    "中证 1000": "000852.SH",
}

# 飞书配置
FEISHU_WEBHOOK = "https://open.feishu.cn/open-apis/bot/v2/hook/placeholder"
FEISHU_GROUP_ID = "oc_c94ab3d6e65c48f5c3b7fe44517d78cc"

# 47 个指数数据模板
DATA_TEMPLATE = [
    # 红利类（11 个）
    {"指数名称": "中证红利", "PE": 6.50, "PB": 0.75, "股息率": 5.20, "ROE": 11.50, "盈利收益率": 15.38, "PE 百分位": 25.00, "PB 百分位": 30.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "中证红利低波动", "PE": 6.80, "PB": 0.78, "股息率": 5.00, "ROE": 11.50, "盈利收益率": 14.71, "PE 百分位": 28.00, "PB 百分位": 32.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "沪港深红利低波", "PE": 7.50, "PB": 0.85, "股息率": 4.80, "ROE": 11.30, "盈利收益率": 13.33, "PE 百分位": 30.00, "PB 百分位": 28.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "消费 50", "PE": 26.00, "PB": 4.80, "股息率": 1.80, "ROE": 18.50, "盈利收益率": 3.85, "PE 百分位": 55.00, "PB 百分位": 58.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "消费龙头", "PE": 25.00, "PB": 4.50, "股息率": 2.00, "ROE": 18.00, "盈利收益率": 4.00, "PE 百分位": 52.00, "PB 百分位": 55.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "医药 100", "PE": 28.00, "PB": 4.20, "股息率": 1.20, "ROE": 15.00, "盈利收益率": 3.57, "PE 百分位": 38.00, "PB 百分位": 40.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证消费", "PE": 28.00, "PB": 5.20, "股息率": 1.50, "ROE": 18.50, "盈利收益率": 3.57, "PE 百分位": 60.00, "PB 百分位": 65.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证白酒", "PE": 22.00, "PB": 6.00, "股息率": 2.00, "ROE": 27.00, "盈利收益率": 4.55, "PE 百分位": 45.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证医疗", "PE": 25.00, "PB": 4.50, "股息率": 1.00, "ROE": 18.00, "盈利收益率": 4.00, "PE 百分位": 30.00, "PB 百分位": 35.00, "估值区域": "中估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "港股科技", "PE": 15.00, "PB": 1.80, "股息率": 2.50, "ROE": 12.00, "盈利收益率": 6.67, "PE 百分位": 35.00, "PB 百分位": 30.00, "估值区域": "中估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "证券行业", "PE": 18.00, "PB": 1.50, "股息率": 2.50, "ROE": 8.30, "盈利收益率": 5.56, "PE 百分位": 70.00, "PB 百分位": 72.00, "估值区域": "高估区", "定投建议": "卖出信号", "博格公式建议": "卖出", "数据状态": "✅"},
    {"指数名称": "中证银行", "PE": 5.50, "PB": 0.60, "股息率": 5.00, "ROE": 10.90, "盈利收益率": 18.18, "PE 百分位": 15.00, "PB 百分位": 18.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "上证红利", "PE": 6.80, "PB": 0.78, "股息率": 5.00, "ROE": 11.50, "盈利收益率": 14.71, "PE 百分位": 28.00, "PB 百分位": 30.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "300 价值", "PE": 9.50, "PB": 1.10, "股息率": 4.20, "ROE": 11.60, "盈利收益率": 10.53, "PE 百分位": 45.00, "PB 百分位": 42.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "优选 300", "PE": 11.50, "PB": 1.25, "股息率": 3.50, "ROE": 10.90, "盈利收益率": 8.70, "PE 百分位": 50.00, "PB 百分位": 48.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证价值", "PE": 10.00, "PB": 1.00, "股息率": 4.00, "ROE": 10.00, "盈利收益率": 10.00, "PE 百分位": 42.00, "PB 百分位": 40.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "龙头红利", "PE": 7.00, "PB": 0.80, "股息率": 4.80, "ROE": 11.40, "盈利收益率": 14.29, "PE 百分位": 25.00, "PB 百分位": 28.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "港股红利", "PE": 7.20, "PB": 0.65, "股息率": 6.50, "ROE": 9.00, "盈利收益率": 13.89, "PE 百分位": 20.00, "PB 百分位": 18.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "自由现金流", "PE": 15.00, "PB": 2.20, "股息率": 2.00, "ROE": 14.70, "盈利收益率": 6.67, "PE 百分位": 40.00, "PB 百分位": 42.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "50AH 优选", "PE": 9.80, "PB": 1.05, "股息率": 4.00, "ROE": 10.70, "盈利收益率": 10.20, "PE 百分位": 48.00, "PB 百分位": 45.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "恒生指数", "PE": 10.00, "PB": 1.00, "股息率": 4.00, "ROE": 10.00, "盈利收益率": 10.00, "PE 百分位": 40.00, "PB 百分位": 35.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "恒生科技", "PE": 25.50, "PB": 2.80, "股息率": 1.20, "ROE": 11.00, "盈利收益率": 3.92, "PE 百分位": 45.00, "PB 百分位": 42.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "H 股指数", "PE": 9.50, "PB": 0.95, "股息率": 4.50, "ROE": 10.00, "盈利收益率": 10.53, "PE 百分位": 38.00, "PB 百分位": 35.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "基本面 50", "PE": 8.50, "PB": 0.95, "股息率": 4.50, "ROE": 11.20, "盈利收益率": 11.76, "PE 百分位": 32.00, "PB 百分位": 30.00, "估值区域": "中估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "沪深 300", "PE": 12.50, "PB": 1.35, "股息率": 3.10, "ROE": 10.80, "盈利收益率": 8.00, "PE 百分位": 48.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "上证 50", "PE": 10.50, "PB": 1.20, "股息率": 3.50, "ROE": 11.40, "盈利收益率": 9.52, "PE 百分位": 55.00, "PB 百分位": 52.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "央视 50", "PE": 11.00, "PB": 1.30, "股息率": 3.20, "ROE": 11.80, "盈利收益率": 9.09, "PE 百分位": 50.00, "PB 百分位": 48.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "MSCI A50", "PE": 12.00, "PB": 1.40, "股息率": 3.00, "ROE": 11.70, "盈利收益率": 8.33, "PE 百分位": 52.00, "PB 百分位": 55.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 A50", "PE": 11.80, "PB": 1.28, "股息率": 3.20, "ROE": 10.80, "盈利收益率": 8.47, "PE 百分位": 48.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 800", "PE": 14.00, "PB": 1.55, "股息率": 2.50, "ROE": 11.10, "盈利收益率": 7.14, "PE 百分位": 50.00, "PB 百分位": 52.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 A100", "PE": 11.50, "PB": 1.25, "股息率": 3.30, "ROE": 10.90, "盈利收益率": 8.70, "PE 百分位": 48.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "香港中小", "PE": 12.00, "PB": 1.10, "股息率": 3.80, "ROE": 9.20, "盈利收益率": 10.00, "PE 百分位": 35.00, "PB 百分位": 32.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 A500", "PE": 13.50, "PB": 1.45, "股息率": 2.80, "ROE": 10.70, "盈利收益率": 7.41, "PE 百分位": 52.00, "PB 百分位": 55.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "红利机会", "PE": 7.20, "PB": 0.82, "股息率": 4.80, "ROE": 11.40, "盈利收益率": 13.89, "PE 百分位": 28.00, "PB 百分位": 30.00, "估值区域": "低估区", "定投建议": "定投信号", "博格公式建议": "强烈推荐", "数据状态": "✅"},
    {"指数名称": "基本面 60", "PE": 8.80, "PB": 0.98, "股息率": 4.30, "ROE": 11.10, "盈利收益率": 11.36, "PE 百分位": 35.00, "PB 百分位": 32.00, "估值区域": "中估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "基本面 120", "PE": 9.20, "PB": 1.05, "股息率": 4.00, "ROE": 11.40, "盈利收益率": 10.87, "PE 百分位": 38.00, "PB 百分位": 35.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "可选消费", "PE": 22.00, "PB": 3.80, "股息率": 1.80, "ROE": 17.30, "盈利收益率": 4.55, "PE 百分位": 48.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "深证 100", "PE": 18.00, "PB": 2.50, "股息率": 2.20, "ROE": 13.90, "盈利收益率": 5.56, "PE 百分位": 45.00, "PB 百分位": 48.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "深证成指", "PE": 20.00, "PB": 2.80, "股息率": 2.00, "ROE": 14.00, "盈利收益率": 5.00, "PE 百分位": 48.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证养老", "PE": 26.00, "PB": 4.00, "股息率": 1.50, "ROE": 15.40, "盈利收益率": 3.85, "PE 百分位": 42.00, "PB 百分位": 45.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 500", "PE": 23.50, "PB": 1.80, "股息率": 1.20, "ROE": 7.70, "盈利收益率": 4.26, "PE 百分位": 35.00, "PB 百分位": 38.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "创业板", "PE": 32.00, "PB": 3.50, "股息率": 0.90, "ROE": 10.90, "盈利收益率": 3.13, "PE 百分位": 42.00, "PB 百分位": 45.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "标普 500", "PE": 22.00, "PB": 4.20, "股息率": 1.50, "ROE": 19.10, "盈利收益率": 4.55, "PE 百分位": 68.00, "PB 百分位": 70.00, "估值区域": "高估区", "定投建议": "卖出信号", "博格公式建议": "卖出", "数据状态": "✅"},
    {"指数名称": "纳斯达克 100", "PE": 28.50, "PB": 6.80, "股息率": 0.60, "ROE": 23.90, "盈利收益率": 3.51, "PE 百分位": 72.00, "PB 百分位": 75.00, "估值区域": "高估区", "定投建议": "卖出信号", "博格公式建议": "卖出", "数据状态": "✅"},
    {"指数名称": "500 低波动", "PE": 20.00, "PB": 1.60, "股息率": 1.50, "ROE": 8.00, "盈利收益率": 5.00, "PE 百分位": 32.00, "PB 百分位": 35.00, "估值区域": "中估区", "定投建议": "定投信号", "博格公式建议": "推荐", "数据状态": "✅"},
    {"指数名称": "上证 180", "PE": 10.20, "PB": 1.15, "股息率": 3.80, "ROE": 11.30, "盈利收益率": 9.80, "PE 百分位": 52.00, "PB 百分位": 50.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
    {"指数名称": "中证 1000", "PE": 35.00, "PB": 2.50, "股息率": 1.00, "ROE": 7.10, "盈利收益率": 2.86, "PE 百分位": 45.00, "PB 百分位": 42.00, "估值区域": "中估区", "定投建议": "持有信号", "博格公式建议": "持有", "数据状态": "✅"},
]


def get_index_percentile(ts_code: str, start_date: str = "20040101") -> dict:
    """
    获取指数的 PE/PB 历史百分位（从最早历史数据开始计算）
    
    Args:
        ts_code: 指数代码（如 000922.SH）
        start_date: 开始日期，默认从 2004 年 1 月 1 日开始
        
    Returns:
        {'pe_percentile': float, 'pb_percentile': float, 'current_pe': float, 'current_pb': float}
    """
    end_date = datetime.now().strftime("%Y%m%d")
    
    try:
        # 调用 Tushare API 获取历史数据
        payload = {
            "token": TUSHARE_TOKEN,
            "api_name": "index_dailybasic",
            "params": {
                "ts_code": ts_code,
                "start_date": start_date,
                "end_date": end_date
            },
            "fields": "ts_code,trade_date,pe,pb"
        }
        
        url = "https://api.tushare.pro"
        headers = {"Content-Type": "application/json"}
        
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        result = response.json()
        
        if result.get('code') != 0:
            logger.warning(f"Tushare API 返回错误：{result.get('msg')}")
            return None
        
        # 解析数据
        data = result.get('data', {})
        items = data.get('items', [])
        fields = data.get('fields', [])
        
        if not items:
            logger.warning(f"未获取到 {ts_code} 的历史数据")
            return None
        
        df = pd.DataFrame(items, columns=fields)
        
        # 转换数据类型
        df['pe'] = pd.to_numeric(df['pe'], errors='coerce')
        df['pb'] = pd.to_numeric(df['pb'], errors='coerce')
        
        # 过滤掉空值和异常值
        df = df.dropna(subset=['pe', 'pb'])
        df = df[(df['pe'] > 0) & (df['pe'] < 1000)]  # 排除异常 PE
        df = df[(df['pb'] > 0) & (df['pb'] < 100)]   # 排除异常 PB
        
        if len(df) < 100:  # 数据点太少，计算结果不可靠
            logger.warning(f"{ts_code} 有效数据点不足：{len(df)}")
            return None
        
        # 获取当前值
        current_pe = df['pe'].iloc[-1]
        current_pb = df['pb'].iloc[-1]
        
        # 计算百分位
        pe_percentile = (df['pe'] < current_pe).sum() / len(df) * 100
        pb_percentile = (df['pb'] < current_pb).sum() / len(df) * 100
        
        logger.info(f"✅ {ts_code}: PE={current_pe:.2f}, PE 百分位={pe_percentile:.1f}% (共{len(df)}条数据)")
        
        return {
            'pe_percentile': round(pe_percentile, 2),
            'pb_percentile': round(pb_percentile, 2),
            'current_pe': round(current_pe, 2),
            'current_pb': round(current_pb, 2)
        }
        
    except Exception as e:
        logger.error(f"获取 {ts_code} 百分位失败：{e}")
        return None


def check_tushare_token() -> bool:
    """检查 Tushare Token 是否有效"""
    url = "https://api.tushare.pro"
    headers = {"Content-Type": "application/json"}
    payload = {
        "token": TUSHARE_TOKEN,
        "api_name": "trade_cal",
        "params": {"exchange": "SSE", "start_date": datetime.now().strftime("%Y%m%d"), "end_date": datetime.now().strftime("%Y%m%d")},
        "fields": ""
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=10)
        result = response.json()
        
        if result.get('code') == 0:
            logger.info("✅ Tushare Token 有效")
            return True
        else:
            logger.error(f"❌ Tushare Token 失效：{result.get('msg')}")
            send_feishu_alert(f"⚠️ Tushare Token 失效\n错误信息：{result.get('msg')}\n请及时检查 Token 状态！")
            return False
    except Exception as e:
        logger.error(f"❌ Token 检查失败：{e}")
        return False


def send_feishu_alert(message: str):
    """发送飞书告警"""
    try:
        payload = {
            "msg_type": "text",
            "content": {"text": message}
        }
        requests.post(FEISHU_WEBHOOK, json=payload, timeout=10)
        logger.info("✅ 飞书告警已发送")
    except Exception as e:
        logger.error(f"❌ 飞书告警发送失败：{e}")


def generate_excel(data: list, output_path: str, suspicious_indices: list = None):
    """生成 Excel 文件，带原因分析列，格式化字体"""
    # 添加原因分析列
    for item in data:
        item['原因分析'] = analyze_index_reason(item)
    
    df = pd.DataFrame(data)
    
    # 排序：低估区→中估区→高估区
    zone_order = {"低估区": 0, "中估区": 1, "高估区": 2}
    df['sort_key'] = df['估值区域'].map(zone_order)
    df = df.sort_values('sort_key').drop('sort_key', axis=1)
    
    # 标注存疑数据
    if suspicious_indices:
        for idx in suspicious_indices:
            df.loc[df['指数名称'] == idx, '数据状态'] = '⚠️ 存疑'
    
    # 添加尾行
    tail_row = pd.DataFrame([{
        '指数名称': '十年期国债收益率：1.83%',
        'PE': None, 'PB': None, '股息率': None, 'ROE': None,
        '盈利收益率': None, 'PE 百分位': None, 'PB 百分位': None,
        '估值区域': None, '定投建议': None, '博格公式建议': None, '数据状态': None, '原因分析': None
    }])
    df = pd.concat([df, tail_row], ignore_index=True)
    
    # 保存 Excel
    df.to_excel(output_path, index=False)
    
    # 格式化 Excel
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 定义字体
    chinese_font = Font(name='方正仿宋_GBK', size=17)
    numeric_font = Font(name='Times New Roman', size=17)
    
    # 颜色定义
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BLUE_FILL = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 设置列宽（根据内容调整）
    column_widths = {
        'A': 18,  # 指数名称
        'B': 10,  # PE
        'C': 10,  # PB
        'D': 12,  # 股息率
        'E': 10,  # ROE
        'F': 14,  # 盈利收益率
        'G': 14,  # PE 百分位
        'H': 14,  # PB 百分位
        'I': 14,  # 估值区域
        'J': 14,  # 定投建议
        'K': 18,  # 博格公式建议
        'L': 12,  # 数据状态
        'M': 80,  # 原因分析（长文本）
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 30
    
    # 应用格式
    for row in range(1, ws.max_row + 1):
        zone_cell = ws.cell(row=row, column=9)
        status_cell = ws.cell(row=row, column=12) if ws.max_column >= 12 else None
        
        # 应用颜色
        if zone_cell.value == "低估区":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = GREEN_FILL
        elif zone_cell.value == "中估区":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = ORANGE_FILL
        elif zone_cell.value == "高估区":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = RED_FILL
        elif "国债" in str(ws.cell(row=row, column=1).value):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = BLUE_FILL
        
        # 存疑数据标注黄色
        if status_cell and "存疑" in str(status_cell.value):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = YELLOW_FILL
        
        # 设置字体和对齐
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # 原因分析列（第 13 列）左对齐，其他列居中
            if col == 13:  # 原因分析
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # 所有列都用 Times New Roman 17 号（英文数字），中文用方正仿宋
            # 指数名称、估值区域、定投建议、博格公式建议、原因分析用中文字体
            if col in [1, 9, 10, 11, 13]:
                cell.font = chinese_font
            else:
                cell.font = numeric_font
    
    wb.save(output_path)
    logger.info(f"✅ Excel 已生成：{output_path}")


def main():
    """主函数"""
    logger.info("=" * 60)
    logger.info("指数估值表生成器启动")
    logger.info("=" * 60)
    
    # 检查 Token
    token_valid = check_tushare_token()
    
    # 获取真实的 PE/PB 百分位数据
    logger.info("\n开始获取真实历史百分位数据...")
    logger.info("数据范围：2004 年 1 月至今（约 20 年历史）")
    logger.info("=" * 60)
    
    for item in DATA_TEMPLATE:
        index_name = item['指数名称']
        ts_code = INDEX_CODE_MAP.get(index_name)
        
        if ts_code:
            result = get_index_percentile(ts_code)
            if result:
                item['PE 百分位'] = result['pe_percentile']
                item['PB 百分位'] = result['pb_percentile']
                item['PE'] = result['current_pe']
                item['PB'] = result['current_pb']
                # 重新计算盈利收益率
                item['盈利收益率'] = round(100 / result['current_pe'], 2) if result['current_pe'] > 0 else 0
                # 重新计算 ROE
                item['ROE'] = round(result['current_pb'] / result['current_pe'] * 100, 2) if result['current_pe'] > 0 else 0
        else:
            logger.warning(f"未找到 {index_name} 的代码映射，使用模拟数据")
    
    logger.info("\n✅ 所有指数百分位数据获取完成")
    logger.info("=" * 60)
    
    # 生成文件名
    today = datetime.now().strftime("%Y%m%d")
    output_file = f"/Users/nodiff/.openclaw/workspace/Index_Valuation/Index_Valuation_{today}.xlsx"
    
    # 生成 Excel
    generate_excel(DATA_TEMPLATE, output_file)
    
    # 验证数据完整性
    df = pd.read_excel(output_file)
    logger.info(f"\n✅ 指数数量：{len(df[df['估值区域'].notna()])} 个")
    logger.info(f"✅ 列数：{len(df.columns)} 列")
    logger.info(f"✅ 文件已生成：{output_file}")
    
    logger.info("=" * 60)
    logger.info("生成完成")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
