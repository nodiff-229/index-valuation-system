"""
指数估值 Excel 报告生成模块

生成格式化的 Excel 报告，包含指数估值数据和投资建议。
支持：
- 博格公式建议列
- 按估值区域自动设置行背景颜色
- 按类别分别标注（稳定行业用盈利收益率法，成长行业用博格公式法，周期行业用 PB 百分位法）
"""

import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# 常量定义
# ============================================================================

DEFAULT_OUTPUT_DIR = Path.home() / ".openclaw" / "workspace" / "Index_Valuation"

VALUATION_THRESHOLDS = {
    "low": 30,
    "high": 70,
}

COLORS = {
    "header": "4472C4",
    "undervalued": "C6EFCE",
    "fair_value": "FFEB9C",
    "overvalued": "FFC7CE",
    "bond_yield": "BDD7EE",
    "white": "FFFFFF",
    "strong_buy": "006600",  # 深绿色 - 强烈推荐
    "buy": "00AA00",        # 绿色 - 推荐
    "hold": "FF9900",       # 橙色 - 持有
    "sell": "FF0000",       # 红色 - 卖出
}

COLUMN_WIDTHS = {
    "指数名称": 15,
    "PE": 10,
    "PB": 10,
    "股息率": 10,
    "ROE": 10,
    "盈利收益率": 12,
    "PE 百分位": 10,
    "PB 百分位": 10,
    "估值区域": 10,
    "定投建议": 12,
    "博格公式建议": 15,
}


# ============================================================================
# 数据类定义
# ============================================================================

@dataclass
class IndexValuationData:
    """指数估值数据（用于 Excel 生成）"""
    name: str
    pe: float
    pb: float
    dividend_yield: float
    roe: float
    earnings_yield: float
    pe_percentile: float
    pb_percentile: float
    valuation_zone: str
    investment_advice: str
    burgess_advice: str           # 博格公式建议
    burgess_color: str            # 博格公式颜色
    category: str = ""            # 指数类别
    expected_return: float = 0.0  # 博格公式预期收益率


# ============================================================================
# Excel 生成器类
# ============================================================================

class ExcelGenerator:
    """Excel 报告生成器"""

    def __init__(self, output_dir: Optional[Path] = None):
        self.output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self._ensure_output_dir()
        logger.info(f"Excel 生成器初始化完成，输出目录：{self.output_dir}")

    def _ensure_output_dir(self) -> None:
        """确保输出目录存在"""
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_valuation_zone(self, pe_percentile: float) -> str:
        """根据 PE 百分位判断估值区域"""
        if pe_percentile < VALUATION_THRESHOLDS["low"]:
            return "低估区"
        elif pe_percentile > VALUATION_THRESHOLDS["high"]:
            return "高估区"
        else:
            return "中估区"

    def _get_investment_advice(self, pe_percentile: float, earnings_yield: float, bond_yield: float) -> str:
        """根据估值和格雷厄姆策略给出定投建议"""
        buy_threshold = 2 * bond_yield
        sell_threshold = bond_yield

        if earnings_yield > buy_threshold:
            return "定投买入"
        elif earnings_yield <= sell_threshold:
            return "分批卖出"
        else:
            if pe_percentile < VALUATION_THRESHOLDS["low"]:
                return "定投买入"
            elif pe_percentile > VALUATION_THRESHOLDS["high"]:
                return "分批卖出"
            else:
                return "继续持有"

    def _calculate_burgess_advice(
        self,
        category: str,
        dividend_yield: float,
        pe: float,
        historical_pe: float,
        roe: float,
        bond_yield: float
    ) -> tuple[str, str, float]:
        """计算博格公式建议
        
        博格公式：预期收益率 = 初始股息率 + 盈利增长率 + PE 变化率
        
        Returns:
            (建议文本，颜色，预期收益率)
        """
        # 稳定行业（消费/医药/红利）：使用盈利收益率法
        stable_categories = ["消费", "医药", "红利"]
        
        # 成长行业（科技）：使用博格公式法
        growth_categories = ["科技"]
        
        # 周期行业（证券/银行）：使用 PB 百分位法
        cyclical_categories = ["金融"]
        
        if category in stable_categories:
            # 盈利收益率法
            earnings_yield = (1 / pe * 100) if pe > 0 else 0
            if earnings_yield > 2 * bond_yield:
                return "盈利收益率法：买入", "绿色", earnings_yield
            elif earnings_yield <= bond_yield:
                return "盈利收益率法：卖出", "红色", earnings_yield
            else:
                return "盈利收益率法：持有", "橙色", earnings_yield
                
        elif category in growth_categories:
            # 博格公式法
            earnings_growth = roe * 0.6  # 假设留存收益率 60%
            pe_change = (historical_pe - pe) / pe * 100 / 10  # 假设 10 年回归
            expected_return = dividend_yield + earnings_growth + pe_change
            
            if expected_return > 15:
                return f"博格公式：强烈推荐 ({expected_return:.1f}%)", "深绿色", expected_return
            elif expected_return > 10:
                return f"博格公式：推荐 ({expected_return:.1f}%)", "绿色", expected_return
            elif expected_return > 5:
                return f"博格公式：持有 ({expected_return:.1f}%)", "橙色", expected_return
            else:
                return f"博格公式：卖出 ({expected_return:.1f}%)", "红色", expected_return
            
        elif category in cyclical_categories:
            # PB 百分位法（简化用 PE 代替）
            pe_percentile = (pe / historical_pe - 1) * 100 + 50  # 简化估算
            if pe_percentile < 30:
                return "PB 百分位法：买入", "绿色", 10.0
            elif pe_percentile > 70:
                return "PB 百分位法：卖出", "红色", 2.0
            else:
                return "PB 百分位法：持有", "橙色", 6.0
        else:
            # 其他类别：使用综合判断
            earnings_yield = (1 / pe * 100) if pe > 0 else 0
            if earnings_yield > 2 * bond_yield:
                return "综合：买入", "绿色", earnings_yield
            elif earnings_yield <= bond_yield:
                return "综合：卖出", "红色", earnings_yield
            else:
                return "综合：持有", "橙色", earnings_yield

    def _get_zone_color(self, valuation_zone: str) -> PatternFill:
        """根据估值区域获取背景颜色"""
        color_map = {
            "低估区": COLORS["undervalued"],
            "中估区": COLORS["fair_value"],
            "高估区": COLORS["overvalued"],
        }
        color = color_map.get(valuation_zone, COLORS["white"])
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _get_burgess_color(self, color_name: str) -> PatternFill:
        """根据博格公式颜色获取填充样式"""
        color_map = {
            "深绿色": COLORS["strong_buy"],
            "绿色": COLORS["buy"],
            "橙色": COLORS["hold"],
            "红色": COLORS["sell"],
        }
        color = color_map.get(color_name, COLORS["white"])
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _sort_by_valuation(self, data: list[IndexValuationData]) -> list[IndexValuationData]:
        """按估值排序：低估→中估→高估"""
        zone_order = {"低估区": 0, "中估区": 1, "高估区": 2}
        return sorted(data, key=lambda x: zone_order.get(x.valuation_zone, 1))

    def _create_header_style(self) -> dict:
        """创建表头样式"""
        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(
                start_color=COLORS["header"],
                end_color=COLORS["header"],
                fill_type="solid"
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _create_cell_style(self, valuation_zone: str) -> dict:
        """创建单元格样式"""
        return {
            "fill": self._get_zone_color(valuation_zone),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _create_burgess_cell_style(self, color_name: str) -> dict:
        """创建博格公式单元格样式"""
        return {
            "fill": self._get_burgess_color(color_name),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _create_bond_yield_style(self) -> dict:
        """创建国债收益率行样式"""
        return {
            "fill": PatternFill(
                start_color=COLORS["bond_yield"],
                end_color=COLORS["bond_yield"],
                fill_type="solid"
            ),
            "font": Font(bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _apply_style(self, cell, style: dict) -> None:
        """应用样式到单元格"""
        for attr, value in style.items():
            setattr(cell, attr, value)

    def generate(
        self,
        data: list[IndexValuationData],
        bond_yield: float = 1.83,
        filename: Optional[str] = None
    ) -> Path:
        """生成 Excel 报告"""
        if not filename:
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"Index_Valuation_{date_str}.xlsx"

        output_path = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "指数估值"

        # 表头（新增博格公式建议列）
        headers = ["指数名称", "PE", "PB", "股息率", "ROE", "盈利收益率",
                   "PE 百分位", "PB 百分位", "估值区域", "定投建议", "博格公式建议"]

        header_style = self._create_header_style()
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, header_style)

        for col, header in enumerate(headers, start=1):
            width = COLUMN_WIDTHS.get(header, 10)
            ws.column_dimensions[get_column_letter(col)].width = width

        sorted_data = self._sort_by_valuation(data)

        # 写入数据
        for row_idx, item in enumerate(sorted_data, start=2):
            row_data = [
                item.name,
                item.pe,
                item.pb,
                f"{item.dividend_yield}%",
                f"{item.roe}%",
                f"{item.earnings_yield}%",
                f"{item.pe_percentile}%",
                f"{item.pb_percentile}%",
                item.valuation_zone,
                item.investment_advice,
                item.burgess_advice,
            ]

            # 前 9 列使用估值区域颜色
            cell_style = self._create_cell_style(item.valuation_zone)
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col, value=row_data[col - 1])
                self._apply_style(cell, cell_style)
            
            # 第 10 列（定投建议）使用估值区域颜色
            cell = ws.cell(row=row_idx, column=10, value=row_data[9])
            self._apply_style(cell, cell_style)
            
            # 第 11 列（博格公式建议）使用博格公式颜色
            burgess_style = self._create_burgess_cell_style(item.burgess_color)
            cell = ws.cell(row=row_idx, column=11, value=row_data[10])
            self._apply_style(cell, burgess_style)

        # 添加国债收益率行
        bond_row = len(sorted_data) + 2
        bond_style = self._create_bond_yield_style()

        ws.merge_cells(start_row=bond_row, start_column=1, end_row=bond_row, end_column=6)
        cell = ws.cell(row=bond_row, column=1, value=f"十年期国债收益率：{bond_yield}%")
        self._apply_style(cell, bond_style)

        for col in range(2, 12):
            cell = ws.cell(row=bond_row, column=col)
            self._apply_style(cell, bond_style)

        ws.freeze_panes = "A2"

        wb.save(output_path)
        logger.info(f"Excel 报告已生成：{output_path}")

        return output_path

    def generate_from_raw_data(
        self,
        raw_data: list[dict],
        bond_yield: float = 1.83,
        filename: Optional[str] = None
    ) -> Path:
        """从原始数据字典生成 Excel 报告"""
        data = []
        for item in raw_data:
            pe_percentile = item.get("pe_percentile", 0)
            earnings_yield = item.get("earnings_yield", 0)
            category = item.get("category", "")
            
            valuation_zone = self._get_valuation_zone(pe_percentile)
            investment_advice = self._get_investment_advice(
                pe_percentile, earnings_yield, bond_yield
            )
            
            # 计算博格公式建议
            historical_pe = item.get("historical_pe", item.get("pe", 15))
            burgess_advice, burgess_color, expected_return = self._calculate_burgess_advice(
                category=category,
                dividend_yield=item.get("dividend_yield", 0),
                pe=item.get("pe", 15),
                historical_pe=historical_pe,
                roe=item.get("roe", 10),
                bond_yield=bond_yield
            )

            data.append(IndexValuationData(
                name=item.get("name", ""),
                pe=item.get("pe", 0),
                pb=item.get("pb", 0),
                dividend_yield=item.get("dividend_yield", 0),
                roe=item.get("roe", 0),
                earnings_yield=earnings_yield,
                pe_percentile=pe_percentile,
                pb_percentile=item.get("pb_percentile", 0),
                valuation_zone=valuation_zone,
                investment_advice=investment_advice,
                burgess_advice=burgess_advice,
                burgess_color=burgess_color,
                category=category,
                expected_return=expected_return,
            ))

        return self.generate(data, bond_yield, filename)


# ============================================================================
# 辅助函数
# ============================================================================

def create_sample_data() -> list[IndexValuationData]:
    """创建示例数据（用于测试）"""
    return [
        IndexValuationData(
            name="中证红利",
            pe=7.5,
            pb=0.85,
            dividend_yield=5.2,
            roe=11.3,
            earnings_yield=13.3,
            pe_percentile=15.5,
            pb_percentile=12.3,
            valuation_zone="低估区",
            investment_advice="定投买入",
            burgess_advice="盈利收益率法：买入",
            burgess_color="绿色",
            category="红利",
        ),
        IndexValuationData(
            name="沪深 300",
            pe=12.8,
            pb=1.35,
            dividend_yield=2.8,
            roe=10.5,
            earnings_yield=7.8,
            pe_percentile=45.2,
            pb_percentile=38.6,
            valuation_zone="中估区",
            investment_advice="继续持有",
            burgess_advice="综合：持有",
            burgess_color="橙色",
            category="宽基",
        ),
        IndexValuationData(
            name="科创 50",
            pe=45.6,
            pb=4.2,
            dividend_yield=0.8,
            roe=9.2,
            earnings_yield=2.2,
            pe_percentile=75.8,
            pb_percentile=82.3,
            valuation_zone="高估区",
            investment_advice="分批卖出",
            burgess_advice="博格公式：卖出 (3.5%)",
            burgess_color="红色",
            category="科技",
        ),
        IndexValuationData(
            name="纳斯达克 100",
            pe=28.5,
            pb=6.8,
            dividend_yield=0.6,
            roe=23.8,
            earnings_yield=3.5,
            pe_percentile=85.2,
            pb_percentile=88.5,
            valuation_zone="高估区",
            investment_advice="分批卖出",
            burgess_advice="博格公式：持有 (8.2%)",
            burgess_color="橙色",
            category="美股",
        ),
    ]


# ============================================================================
# 主函数
# ============================================================================

def main():
    """测试函数"""
    print("\n" + "=" * 60)
    print("Excel 报告生成器测试")
    print("=" * 60)

    generator = ExcelGenerator()
    sample_data = create_sample_data()
    output_path = generator.generate(sample_data, bond_yield=1.83)

    print(f"\n报告已生成：{output_path}")
    print("=" * 60)

    print("\n测试从原始数据生成...")
    raw_data = [
        {
            "name": "创业板指",
            "pe": 35.2,
            "pb": 4.5,
            "dividend_yield": 1.2,
            "roe": 12.8,
            "earnings_yield": 2.8,
            "pe_percentile": 55.5,
            "pb_percentile": 48.2,
            "category": "科技",
        },
        {
            "name": "上证 50",
            "pe": 9.8,
            "pb": 1.1,
            "dividend_yield": 3.5,
            "roe": 11.2,
            "earnings_yield": 10.2,
            "pe_percentile": 22.3,
            "pb_percentile": 18.5,
            "category": "宽基",
        },
    ]

    output_path2 = generator.generate_from_raw_data(raw_data, filename="test_raw_data.xlsx")
    print(f"原始数据报告已生成：{output_path2}")

    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)


if __name__ == "__main__":
    main()
